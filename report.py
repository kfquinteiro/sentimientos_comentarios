"""Gera o relatório final em XLSX: dados consolidados + análise de
sentimento + gráficos (donut general, por red, en el tiempo)."""
import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import BarChart, DoughnutChart, LineChart, Reference
from openpyxl.utils import get_column_letter

SENTIMENT_ORDER = ["Positivo", "Neutral", "Negativo"]

COLUMN_LABELS = {
    "marca": "Marca",
    "red": "Red",
    "link_post": "Link del post",
    "fecha_post": "Fecha de publicación",
    "autor": "Autor",
    "autor_id": "ID del autor",
    "fecha_comentario": "Fecha del comentario",
    "likes": "Likes",
    "comentario": "Comentario",
    "sentimiento": "Sentimiento",
    "confianza": "Confianza",
}


def _clean(value):
    if pd.isna(value):
        return None
    if hasattr(value, "to_pydatetime"):
        return value.to_pydatetime()
    return value


def _write_dataframe(ws, df, start_row):
    """Escreve df (cabeçalho + dados) a partir de start_row. Retorna a última linha usada."""
    for c_idx, col in enumerate(df.columns, start=1):
        ws.cell(row=start_row, column=c_idx, value=str(col))
    for r_offset, row in enumerate(df.itertuples(index=False), start=1):
        for c_idx, value in enumerate(row, start=1):
            ws.cell(row=start_row + r_offset, column=c_idx, value=_clean(value))
    return start_row + len(df)


def build_report(df, output_path):
    """df: DataFrame consolidado (salida de consolidate.build_consolidated)
    con las columnas adicionales 'sentimiento' y 'confianza'."""
    df = df.copy()

    wb = Workbook()

    # --- Hoja 1: datos ---
    ws_data = wb.active
    ws_data.title = "Comentarios"
    data_df = df.rename(columns=COLUMN_LABELS)
    _write_dataframe(ws_data, data_df, start_row=1)
    for c_idx in range(1, len(data_df.columns) + 1):
        ws_data.column_dimensions[get_column_letter(c_idx)].width = 22

    # --- Hoja 2: resumen + gráficos ---
    ws = wb.create_sheet("Resumen")

    # Tabla 1: sentimiento general
    overall = (
        df["sentimiento"].value_counts()
        .reindex(SENTIMENT_ORDER, fill_value=0)
        .reset_index()
    )
    overall.columns = ["Sentimiento", "Cantidad"]
    overall_start = 1
    overall_end = _write_dataframe(ws, overall, overall_start)

    donut = DoughnutChart()
    donut.title = "Sentimiento general"
    donut.add_data(Reference(ws, min_col=2, min_row=overall_start, max_row=overall_end), titles_from_data=True)
    donut.set_categories(Reference(ws, min_col=1, min_row=overall_start + 1, max_row=overall_end))
    ws.add_chart(donut, "E1")

    # Tabla 2: sentimiento por red
    by_network = (
        df.groupby("red")["sentimiento"]
        .value_counts().unstack(fill_value=0)
        .reindex(columns=SENTIMENT_ORDER, fill_value=0)
        .reset_index()
        .rename(columns={"red": "Red"})
    )
    by_network_start = overall_end + 2
    by_network_end = _write_dataframe(ws, by_network, by_network_start)

    bar = BarChart()
    bar.type = "col"
    bar.grouping = "clustered"
    bar.title = "Sentimiento por red"
    bar.add_data(
        Reference(ws, min_col=2, max_col=1 + len(SENTIMENT_ORDER), min_row=by_network_start, max_row=by_network_end),
        titles_from_data=True,
    )
    bar.set_categories(Reference(ws, min_col=1, min_row=by_network_start + 1, max_row=by_network_end))
    ws.add_chart(bar, "E18")

    # Tabla 3: sentimiento en el tiempo (por mes del comentario)
    over_time_start = by_network_end + 2
    with_date = df.dropna(subset=["fecha_comentario"]).copy()
    if not with_date.empty:
        with_date["mes"] = with_date["fecha_comentario"].dt.to_period("M").astype(str)
        over_time = (
            with_date.groupby("mes")["sentimiento"]
            .value_counts().unstack(fill_value=0)
            .reindex(columns=SENTIMENT_ORDER, fill_value=0)
            .reset_index()
            .sort_values("mes")
        )
        over_time = over_time.rename(columns={"mes": "Mes"})
    else:
        over_time = pd.DataFrame(columns=["Mes"] + SENTIMENT_ORDER)
    over_time_end = _write_dataframe(ws, over_time, over_time_start)

    if over_time_end > over_time_start:
        line = LineChart()
        line.title = "Sentimiento en el tiempo"
        line.add_data(
            Reference(ws, min_col=2, max_col=1 + len(SENTIMENT_ORDER), min_row=over_time_start, max_row=over_time_end),
            titles_from_data=True,
        )
        line.set_categories(Reference(ws, min_col=1, min_row=over_time_start + 1, max_row=over_time_end))
        ws.add_chart(line, "E36")

    for c_idx in range(1, 6):
        ws.column_dimensions[get_column_letter(c_idx)].width = 14

    wb.save(output_path)
